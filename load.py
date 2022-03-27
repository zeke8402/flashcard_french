from openpyxl import Workbook, load_workbook

# 'en' is english
# 'fr' is french
def translate_text(target, text):
    """Translates text into the target language.

    Target must be an ISO 639-1 language code.
    See https://g.co/cloud/translate/v2/translate-reference#supported_languages
    """
    import six
    from google.cloud import translate_v2 as translate

    translate_client = translate.Client()

    if isinstance(text, six.binary_type):
        text = text.decode("utf-8")

    # Text can also be a sequence of strings, in which case this method
    # will return a sequence of results for each text.
    result = translate_client.translate(text, target_language=target)
    return result["translatedText"]


wb = Workbook()
src_wb = load_workbook('french.xlsx')
ws = src_wb.active
for row in ws.iter_rows(min_row=501, max_col=1, max_row=1000):
    for cell in row:
        print(cell.value)
        ws.cell(row=cell.row, column=2).value = translate_text('en', cell.value)
        src_wb.save('french.xlsx')

